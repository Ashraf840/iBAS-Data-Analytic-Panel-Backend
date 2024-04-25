from django.shortcuts import render, HttpResponse
from qaDatasetApp.models import qa_dataset as qadm
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
import pandas as pd
import os
from .models import SuggestiveQuestions
from .serializers import SuggestiveQuestionsSerializer
from django.db import connection
import requests
from openpyxl import load_workbook


@api_view(['POST'])
def addToDataset(request):
    if request.method == "POST":
        print(request.data)
        record_id = request.data
        return Response({'msg': 'Added to dataset'}, status=status.HTTP_200_OK)
        
        try:
            # Query the database to retrieve the specific record by its ID
            record = qadm.QADataset.objects.get(pk=record_id)
            
            # Extract the fields you want from the record
            bangla_ques = record.bangla_ques
            english_ques = record.english_ques
            tranliterated_ques = record.transliterated_ques
            bangla_ans = record.bangla_ans
            english_ans = record.english_ans

            # You can return these fields in the response
            data = {
                'bangla_ques': bangla_ques,
                'english_ques': english_ques,
                'tranliterated_ques': tranliterated_ques,
                'bangla_ans': bangla_ans,
                'english_ans': english_ans,
            }
            print("data:",data)

            new_update_dataset = '/media/robin/Documents/PersonalWorks/ibas_project/source/ibas_final_dataset.xlsx'
            # new_update_dataset = '/home/tanjim/workstation/ibas-project/ibas-chat-operator-chatbot/data/ibas_final_dataset.xlsx'
            existing_df = pd.read_excel(new_update_dataset, sheet_name='Sheet1', engine='openpyxl')
            data1 = pd.DataFrame({'Questions': [bangla_ques], 'Answers': [bangla_ans]})
            data2 = pd.DataFrame({'Questions': [tranliterated_ques], 'Answers': [bangla_ans]})
            data3 = pd.DataFrame({'Questions': [english_ques], 'Answers': [english_ans]})

            updated_df = pd.concat([existing_df, data1, data2, data3], ignore_index=True)
            updated_df.to_excel(new_update_dataset, index=False, engine='openpyxl')

            record.flags = True
            record.save()

            return Response(data, status=status.HTTP_200_OK)

        except qadm.QADataset.DoesNotExist:
            return Response({'msg': 'Record not found'}, status=status.HTTP_404_NOT_FOUND)
    
    return Response({'msg': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
def pushSuggestiveQA(request):
    if request.method == "GET":
        print("Insert new paraphrased data into the DB!")

        # return Response({'msg': 'Successful'}, status=status.HTTP_200_OK)
        
        xlsx_file_path = '/media/robin/Documents/PersonalWorks/ibas_project/source/paraphrased_texts.xlsx'
        # new_update_dataset = '/home/tanjim/workstation/ibas-project/ibas-chat-operator-chatbot/data/ibas_final_dataset.xlsx'
        if os.path.exists(xlsx_file_path):
            try:
                wb = load_workbook(xlsx_file_path)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Assuming the structure of your Excel file:
                    # Column 1: Field 1, Column 2: Field 2, etc.

                    field1 = row[0]
                    # print(field1)
                    # Map other fields as needed

                    # Check if there are already exisiting same paraphrased text in the DB, only insert the new ones.
                    instances = SuggestiveQuestions.objects.filter(text=field1)
                    if instances.exists():
                        print("Found an exisiting paraphrased texts in the DB")
                        # instance = instances.first()
                    else:
                        # Create an instance of your model
                        instance = SuggestiveQuestions(text=field1)
                        # Save the instance to the database
                        instance.save()


                # df = pd.read_excel(xlsx_file_path, sheet_name='Sheet1')
                # for index, row in df.iterrows():
                #     text=row['ParaphrasedText']
                #     # questions = SuggestiveQuestions.objects.filter(marked_for_removal=False)
                #     # serializer = SuggestiveQuestionsSerializer(questions, many=True)
                #     entry = SuggestiveQuestions(text=text,)
                #     entry.save()
                return Response({'msg': 'Successful'}, status=status.HTTP_200_OK)
            except Exception as e:
                return HttpResponse(f'Error importing questions: {str(e)}')
        return Response({'msg': 'Get all suggestive qna'}, status=status.HTTP_200_OK)
    if request.method == "POST":
        pass

# Suggestive QnA
@api_view(['GET', 'POST'])
def suggestiveQA(request):
    if request.method == "GET":
        try:
            offset = int(request.query_params.get('offset', 0))
            limit = int(request.query_params.get('limit', 0))
            text = request.query_params.get('searchText', '')
            count = SuggestiveQuestions.objects.filter(marked_for_removal=False).count()
            cursor = connection.cursor()
            query = """
            SELECT * FROM public."addToDataset_suggestivequestions"
            WHERE text ILIKE %s 
            AND marked_for_removal = false
            ORDER BY id DESC;
            """
            
            if text and not limit:
                cursor.execute(query, ['%' + text + '%'])
                rows = cursor.fetchall()
                count = len(rows)
                print("not limit")
                questions = []
                for row in rows:
                    data = {
                    'id': row[0],
                    'text': row[1],
                    'answer': row[2],
                    'marked_for_removal': row[3],
                    'is_added_to_qa_dataset': row[4]
                    }
                    questions.append(data)
                # questions = SuggestiveQuestions.objects.filter(marked_for_removal=False,text__icontains=text)[offset:offset]
            elif limit and not text:
                questions = SuggestiveQuestions.objects.filter(marked_for_removal=False).order_by('-id')[offset:offset + limit]
            elif limit and text:
                cursor.execute(query, ['%' + text + '%'])
                row = cursor.fetchall()
                count = len(row)
                cursor.execute(query + " OFFSET %s LIMIT %s", ['%' + text + '%']+ [offset, limit])
                rows = cursor.fetchall()
                # count = len(rows)
                print("limit::::::::::", count)

                # print("questions", questions)
                questions = []
                for row in rows:
                    data = {
                    'id': row[0],
                    'text': row[1],
                    'answer': row[2],
                    'marked_for_removal': row[3],
                    'is_added_to_qa_dataset': row[4]
                    }
                    questions.append(data)
                # questions = SuggestiveQuestions.objects.filter(marked_for_removal=False,question__icontains=text)[offset:offset+limit]
                print('questions', questions)
            else:
                questions = SuggestiveQuestions.objects.filter(marked_for_removal=False).order_by('-id')
            serializer = SuggestiveQuestionsSerializer(questions, many=True)
            return Response({'data': serializer.data, 'count': count}, status=status.HTTP_200_OK)
        except Exception as e:
            return HttpResponse(f'Error importing questions: {str(e)}')
    if request.method == "POST":
        pass

@api_view(['GET'])
def genSuggestiveQa(request):
    if request.method == "GET":

        # # TESTING *********************
        # response = {
        #     'status_code': 200
        # }

        # if response['status_code'] == 200:
        #     requests.get('http://127.0.0.1:8082/push-suggestive-qa/')
        #     return Response({
        #         'status_code': 200,
        #         'msg': 'Generate suggestive question QA successfully'
        #         }, status=status.HTTP_200_OK)
        

        print("Invoke generate-suggestive-questions API")     # Data Clustering model to generate base form of augmented question with same meaning

        url = 'http://127.0.0.1:5001/suggestive_ques_gen'  # Data Seggregator API

        response = requests.get(url)

        # # Check if the request was successful (status code 200)
        if response.status_code == 200:
            # Print the response content
            print(response.text)











            # NEED TO BE MOVED TO THE "ibas_data_seggregator" project
            # requests.get('http://127.0.0.1:8082/push-suggestive-qa/')   # Push the new paraphrased queries to db; TODO: Shift from this backend to ibas_data_seggregator project & invoke inside the run.sh file 











            return Response({
                'status_code': 200,
                'msg': 'Generate suggestive question QA successfully'
                }, status=status.HTTP_200_OK)
        else:
            # Print an error message if the request was not successful
            print(f"Failed to fetch data. Status code: {response.status_code}")
        
            return Response({
                'status_code': 500,
                'msg': 'Failed generate suggestive question QA'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)